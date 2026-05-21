from __future__ import annotations

import math
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook


# ------------------------------------------------------
# REQUIRED by your email script
# ------------------------------------------------------
WORKBOOK_PATH = r"C:\Users\hugom\OneDrive\Desktop\Root\Personal\Projects\Portfolio\Portfolio.xlsx"


# ------------------------------------------------------
# CONFIG
# ------------------------------------------------------
PORTFOLIO_SHEET = "portfolio"
PNL_SHEET = "PnL"

DOWNLOAD_PERIOD = "3y"
WINDOW = 252
RISK_FREE_ANNUAL = 0.0

REQ_PORT = {"Symbol", "YF_Ticker", "Shares", "Price"}
REQ_PNL = {"YF_Ticker", "Shares", "Prev_Close", "Last_Close", "Daily_PnL_EUR", "Start_Of_Year_Close", "YTD_Return_Pct"}


def _norm(s: str) -> str:
    return (s or "").strip()


def _to_float(x, default=0.0) -> float:
    try:
        if x is None:
            return default
        if isinstance(x, str) and x.strip() == "":
            return default
        if isinstance(x, str):
            x = x.replace(",", "")
        return float(x)
    except Exception:
        return default


def _find_header_row(ws, required: set[str], max_scan_rows: int = 30) -> Tuple[int, Dict[str, int]]:
    """
    Returns (header_row_1_based, header_map[name] = col_1_based)
    """
    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        headers = [_norm(str(v)) if v is not None else "" for v in row_vals]
        if not any(headers):
            continue
        if required.issubset(set(headers)):
            hm = {h: headers.index(h) + 1 for h in headers if h}  # first occurrence
            return r, hm
    raise RuntimeError(f"Could not find required headers: {sorted(required)}")


def _series_map(dl: pd.DataFrame, tickers: List[str], field: str) -> Dict[str, pd.Series]:
    out: Dict[str, pd.Series] = {}
    if dl is None or dl.empty:
        return out

    if isinstance(dl.columns, pd.MultiIndex):
        # yfinance typically returns columns like (ticker, field)
        for t in tickers:
            key = (t, field)
            if key in dl.columns:
                out[t] = dl[key].dropna()
    else:
        # single ticker case
        if field in dl.columns and len(tickers) == 1:
            out[tickers[0]] = dl[field].dropna()

    return out


def ann_vol_and_sharpe(rets: pd.Series, window: int, rf_annual: float) -> Tuple[Optional[float], Optional[float]]:
    r = rets.dropna()
    if r.shape[0] < 2:
        return None, None
    r = r.iloc[-window:] if r.shape[0] >= window else r
    if r.shape[0] < 2:
        return None, None

    vol = float(r.std(ddof=1)) * math.sqrt(252.0)
    if vol == 0.0:
        return vol, None

    mean_annual = float(r.mean()) * 252.0
    sharpe = (mean_annual - rf_annual) / vol
    return vol, float(sharpe)


def main(xlsx_path: str = WORKBOOK_PATH, save: bool = True) -> None:
    """
    - Reads Portfolio sheet positions (Symbol, YF_Ticker, Shares, Price)
    - Updates Portfolio.Price with latest Close
    - Updates PnL sheet per-ticker metrics (matching by YF_Ticker)
    - Prints summary lines parsed by your email script
    """
    wb = load_workbook(xlsx_path)
    if PORTFOLIO_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{PORTFOLIO_SHEET}' in workbook.")
    if PNL_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{PNL_SHEET}' in workbook.")

    ws_port = wb[PORTFOLIO_SHEET]
    ws_pnl = wb[PNL_SHEET]

    # --- headers ---
    port_hdr_row, port_hm = _find_header_row(ws_port, REQ_PORT)
    pnl_hdr_row, pnl_hm = _find_header_row(ws_pnl, REQ_PNL)

    # --- read Portfolio positions ---
    positions: List[Dict[str, object]] = []
    cash_eur = 0.0

    r = port_hdr_row + 1
    while r <= ws_port.max_row:
        sym = _norm(str(ws_port.cell(r, port_hm["Symbol"]).value or ""))
        if not sym:
            r += 1
            continue
        if sym.strip().lower() == "total":
            break

        sym_l = sym.strip().lower()
        if sym_l == "cash":
            # Prefer Market_Value if present; otherwise price*shares
            mv_col = port_hm.get("Market_Value")
            if mv_col:
                cash_eur = _to_float(ws_port.cell(r, mv_col).value, 0.0)
            else:
                px = _to_float(ws_port.cell(r, port_hm["Price"]).value, 0.0)
                sh = _to_float(ws_port.cell(r, port_hm["Shares"]).value, 0.0)
                cash_eur = px * sh
            r += 1
            continue

        tkr = _norm(str(ws_port.cell(r, port_hm["YF_Ticker"]).value or ""))
        sh = _to_float(ws_port.cell(r, port_hm["Shares"]).value, 0.0)

        if tkr and tkr.lower() != "nan" and sh != 0.0:
            positions.append({"row": r, "ticker": tkr, "shares": sh})
        r += 1

    if not positions:
        raise RuntimeError("No non-zero positions found on Portfolio sheet (excluding Cash).")

    tickers = [p["ticker"] for p in positions]

    # --- download data once ---
    dl = yf.download(
        tickers=tickers,
        period=DOWNLOAD_PERIOD,
        interval="1d",
        auto_adjust=False,
        group_by="ticker",
        threads=True,
        progress=False,
    )

    close_map = _series_map(dl, tickers, "Close")
    adj_map = _series_map(dl, tickers, "Adj Close")
    for t in tickers:
        if t not in adj_map:
            adj_map[t] = close_map.get(t, pd.Series(dtype=float))

    # --- update Portfolio.Price ---
    for p in positions:
        t = p["ticker"]
        s = close_map.get(t)
        if s is not None and not s.empty:
            ws_port.cell(p["row"], port_hm["Price"]).value = float(s.iloc[-1])

    # --- update PnL sheet per ticker (match by YF_Ticker) ---
    # Build a map from PnL ticker -> row
    pnl_row_by_ticker: Dict[str, int] = {}
    rr = pnl_hdr_row + 1
    while rr <= ws_pnl.max_row:
        t = _norm(str(ws_pnl.cell(rr, pnl_hm["YF_Ticker"]).value or ""))
        if t:
            pnl_row_by_ticker[t] = rr
        rr += 1

    today_year = datetime.today().year

    # fill from positions (shares from Portfolio overrides PnL shares if you want)
    for p in positions:
        t = p["ticker"]
        sh = float(p["shares"])
        s = close_map.get(t, pd.Series(dtype=float)).dropna()
        if s.empty:
            continue

        last_close = float(s.iloc[-1])
        prev_close = float(s.iloc[-2]) if len(s) >= 2 else last_close
        daily_pnl = (last_close - prev_close) * sh

        s_y = s[s.index.year == today_year]
        start_of_year = float(s_y.iloc[0]) if not s_y.empty else last_close
        ytd = (last_close / start_of_year - 1.0) * 100.0 if start_of_year else 0.0

        row = pnl_row_by_ticker.get(t)
        if row is None:
            # If ticker not present on PnL sheet, skip (don’t destroy user layout)
            continue

        ws_pnl.cell(row, pnl_hm["Shares"]).value = sh
        ws_pnl.cell(row, pnl_hm["Prev_Close"]).value = prev_close
        ws_pnl.cell(row, pnl_hm["Last_Close"]).value = last_close
        ws_pnl.cell(row, pnl_hm["Daily_PnL_EUR"]).value = daily_pnl
        ws_pnl.cell(row, pnl_hm["Start_Of_Year_Close"]).value = start_of_year
        ws_pnl.cell(row, pnl_hm["YTD_Return_Pct"]).value = ytd

    # --- portfolio-level metrics (printed for email) ---
    # Daily PnL
    portfolio_daily_pnl = 0.0
    for p in positions:
        t = p["ticker"]
        sh = float(p["shares"])
        s = close_map.get(t, pd.Series(dtype=float)).dropna()
        if len(s) < 2:
            continue
        portfolio_daily_pnl += (float(s.iloc[-1]) - float(s.iloc[-2])) * sh

    # YTD portfolio return (same-currency assumption; cash constant)
    total_start = float(cash_eur)
    total_curr = float(cash_eur)
    for p in positions:
        t = p["ticker"]
        sh = float(p["shares"])
        s = close_map.get(t, pd.Series(dtype=float)).dropna()
        if s.empty:
            continue
        s_y = s[s.index.year == today_year]
        start_px = float(s_y.iloc[0]) if not s_y.empty else float(s.iloc[-1])
        last_px = float(s.iloc[-1])
        total_start += start_px * sh
        total_curr += last_px * sh

    ytd_pct = None
    if total_start != 0.0:
        ytd_pct = (total_curr / total_start - 1.0) * 100.0

    # 252d vol + sharpe (adj close)
    adj_df = pd.concat(
        {t: adj_map[t] for t in tickers if t in adj_map and not adj_map[t].empty},
        axis=1,
        join="inner",
    ).dropna(how="any")

    vol_252 = None
    sharpe_252 = None
    if adj_df.shape[0] >= 3:
        shares_map = {p["ticker"]: float(p["shares"]) for p in positions}
        share_vec = pd.Series(shares_map).reindex(adj_df.columns).astype(float)
        port_value = (adj_df * share_vec).sum(axis=1) + float(cash_eur)
        port_rets = port_value.pct_change()
        vol_252, sharpe_252 = ann_vol_and_sharpe(port_rets, WINDOW, float(RISK_FREE_ANNUAL))

    # Save if requested (your daily email run should save)
    if save:
        wb.save(xlsx_path)

    # ---- These printed keys MUST match your email parser ----
    print(f"Portfolio Daily PnL (EUR): {portfolio_daily_pnl:,.2f}")
    print("Portfolio YTD Return (%): " + ("N/A" if ytd_pct is None else f"{ytd_pct:,.4f}"))
    print("Portfolio Vol_252d (ann): " + ("N/A" if vol_252 is None else f"{vol_252:.6f}"))
    print("Portfolio Sharpe_252d:    " + ("N/A" if sharpe_252 is None else f"{sharpe_252:.6f}"))


if __name__ == "__main__":
    main(WORKBOOK_PATH)
