import io
import os
import smtplib
import warnings
from contextlib import redirect_stdout
from datetime import date, datetime
from email.message import EmailMessage
from html import escape as html_escape
from pathlib import Path

from openpyxl import load_workbook
from src.quote import pick_random_quote
from src.mean import make_insult
from src.training import get_today_plan, format_plan_for_email
from src.portfolio import build_portfolio_html

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
    category=UserWarning,
    module=r"openpyxl\.worksheet\._reader",
)

PASSAGES_FOLDER = Path(__file__).resolve().parent / "data" / "passages"
PDF_PATTERN = "Passages - *.pdf"
RECURSIVE = False
GAP_MULTIPLIER = 1.8
TRAINING_PLAN_PATH = Path(__file__).resolve().parent / "data" / "Training_Plan.xlsx"
SATURDAY = 5


def load_dotenv(dotenv_path: Path) -> None:
    if not dotenv_path.exists():
        return
    for line in dotenv_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


def send_email_smtp(
    smtp_host: str,
    smtp_port: int,
    smtp_user: str,
    smtp_pass: str,
    mail_from: str,
    mail_to: str,
    subject: str,
    body_text: str,
    body_html: str | None = None,
    mail_cc: str | list[str] | None = None,
) -> None:
    msg = EmailMessage()
    msg["From"] = mail_from
    msg["To"] = mail_to
    if mail_cc:
        msg["Cc"] = ", ".join(mail_cc) if isinstance(mail_cc, list) else mail_cc
    msg["Subject"] = subject
    msg.set_content(body_text)
    if body_html:
        msg.add_alternative(body_html, subtype="html")
    if smtp_port == 465:
        with smtplib.SMTP_SSL(smtp_host, smtp_port) as s:
            s.login(smtp_user, smtp_pass)
            s.send_message(msg)
    else:
        with smtplib.SMTP(smtp_host, smtp_port) as s:
            s.ehlo()
            s.starttls()
            s.login(smtp_user, smtp_pass)
            s.send_message(msg)


def prettify_quote(text: str) -> str:
    lines = [ln.lstrip().lstrip("> ").rstrip() for ln in text.splitlines()]
    return "\n".join(lines).strip()


def bullet(text: str) -> str:
    return f"• {text}"


def strip_bullet_prefix(s: str) -> str:
    return s[2:] if s.startswith("• ") else s


def _strip_done_prefix(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return s
    sl = s.lower()
    if sl.startswith("done"):
        s = s[4:].lstrip()
        if s.startswith(":") or s.startswith("-"):
            s = s[1:].lstrip()
    return s.strip()


def _norm_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "").replace("_", "")


def _cell_date_to_date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
    return None


def _find_header_row(ws, max_scan_rows: int = 30) -> tuple[int, dict[str, int]] | None:
    required = {"week", "date", "day"}
    known = {"run_km", "runkm", "run_rpe", "runrpe", "run_session", "runsession", "gym", "checklist"}
    for r in range(1, max_scan_rows + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        normed = [_norm_header(str(v)) if v is not None else "" for v in values]
        if not any(normed):
            continue
        header_map: dict[str, int] = {}
        for idx, h in enumerate(normed, start=1):
            if not h:
                continue
            if h == "runkm":
                header_map["run_km"] = idx
            elif h == "runrpe":
                header_map["run_rpe"] = idx
            elif h == "runsession":
                header_map["run_session"] = idx
            elif h in {"week", "date", "day", "gym", "checklist"}:
                header_map[h] = idx
            elif h in known:
                header_map[h] = idx
        if required.issubset(header_map):
            return r, header_map
    return None


def get_today_training_plan_row(plan_path: Path, target: date) -> dict[str, str]:
    if not plan_path.exists():
        return {"_error": f"Training plan not found: {plan_path}"}
    try:
        wb = load_workbook(plan_path, data_only=True, read_only=True)
    except Exception as e:
        return {"_error": f"Could not open training plan: {e}"}
    ws = wb.active
    hdr = _find_header_row(ws)
    if hdr is None:
        return {"_error": "Could not find header row in training plan (expected Week/Date/Day columns)."}
    header_row, hm = hdr
    best_future = None
    found_exact = None

    def get(colname: str, row_idx: int) -> str:
        col = hm.get(colname)
        if not col:
            return ""
        v = ws.cell(row=row_idx, column=col).value
        return "" if v is None else str(v).strip()

    for r in range(header_row + 1, ws.max_row + 1):
        d = _cell_date_to_date(ws.cell(row=r, column=hm["date"]).value)
        if d is None:
            continue
        row = {
            "week": get("week", r),
            "date": d.strftime("%d/%m/%Y"),
            "day": get("day", r),
            "run_km": get("run_km", r),
            "run_rpe": get("run_rpe", r),
            "run_session": get("run_session", r),
            "gym": get("gym", r),
            "checklist": get("checklist", r),
        }
        if d == target:
            found_exact = row
            break
        if d > target:
            if best_future is None or d < best_future[0]:
                best_future = (d, row)
    if found_exact is not None:
        return found_exact
    if best_future is not None:
        d, row = best_future
        row["_note"] = (
            f"No exact match for {target.strftime('%d/%m/%Y')}; "
            f"showing next planned day {d.strftime('%d/%m/%Y')}."
        )
        return row
    return {"_error": f"No matching dates found in plan for/after {target.strftime('%d/%m/%Y')}."}


def format_training_plan_row(row: dict[str, str]) -> list[str]:
    if "_error" in row:
        return [bullet(row["_error"])]
    week = (row.get("week") or "").strip()
    dte = (row.get("date") or "").strip()
    day = (row.get("day") or "").strip()
    run_km = (row.get("run_km") or "").strip()
    run_rpe = (row.get("run_rpe") or "").strip()
    run_session = _strip_done_prefix((row.get("run_session") or "").strip())
    gym = (row.get("gym") or "").strip()
    checklist = (row.get("checklist") or "").strip()
    out: list[str] = []
    title_bits = []
    if week:
        title_bits.append(f"Week {week}")
    if day or dte:
        title_bits.append("—")
        title_bits.append(" ".join(x for x in [day, dte] if x))
    if title_bits:
        out.append(bullet(" ".join(title_bits)))
    run_line = f"Run: {run_km} km" if run_km and run_km not in {"-", ""} else "Run: (no run)"
    if run_rpe and run_rpe not in {"-", ""}:
        run_line += f" (RPE {run_rpe})"
    if run_session and run_session not in {"-", ""}:
        run_line += f" — {run_session}"
    out.append(bullet(run_line))
    if gym and gym not in {"-", ""}:
        out.append(bullet(f"Gym: {gym}"))
    if checklist and checklist not in {"-", ""}:
        out.append(bullet(f"Checklist: {checklist}"))
    note = (row.get("_note") or "").strip()
    if note:
        out.append(bullet(note))
    return out


def main() -> int:
    here = Path(__file__).resolve().parent
    load_dotenv(here / ".env")

    smtp_host = os.environ.get("SMTP_HOST", "")
    smtp_port = int(os.environ.get("SMTP_PORT", "465"))
    smtp_user = os.environ.get("SMTP_USER", "")
    smtp_pass = os.environ.get("SMTP_PASS", "")
    mail_from = os.environ.get("MAIL_FROM", smtp_user)
    mail_to = os.environ.get("MAIL_TO", "")

    missing = [
        k for k in ["SMTP_HOST", "SMTP_PORT", "SMTP_USER", "SMTP_PASS", "MAIL_FROM", "MAIL_TO"]
        if not os.environ.get(k)
    ]
    if missing:
        raise SystemExit(f"Missing settings in .env: {', '.join(missing)}")

    # ---- Workout ----
    plan_row = get_today_plan()
    workout_plain, workout_html = format_plan_for_email(plan_row)
    
    # ---- Portfolio ----
    portfolio_html = build_portfolio_html()
    
    # ---- Weather ----
    location = plan_row.get("location", "dublin") if plan_row else "dublin"
    from src.training import get_weather
    weather_str = get_weather(location)

    # ---- Quote ----
    q = pick_random_quote(
        PASSAGES_FOLDER,
        pattern=PDF_PATTERN,
        recursive=RECURSIVE,
        gap_multiplier=GAP_MULTIPLIER,
    )
    pretty = prettify_quote(q.text)

    # ---- Build email ----
    today_str = date.today().strftime("%a %d %b %Y")
    subject = f"Daily Update: {today_str}"

    # Plain text
    lines: list[str] = []
    lines.append(subject)
    lines.append("")
    lines.append("Portfolio")
    lines.append("See HTML version for portfolio details.")
    lines.append("")
    lines.append("Workout")
    lines.append(workout_plain)
    lines.append("")
    lines.append("Quote")
    lines.append(f"\u201c{pretty}\u201d")
    lines.append(f"— {q.author}, {q.title}" if q.author else f"— {q.title}")
    body_text = "\n".join(lines)

    # HTML
    html_parts: list[str] = []
    html_parts.append("<div style='max-width:600px; margin:0 auto; padding:1rem; font-family:Arial,sans-serif;'>")

    WEATHER_ICONS = {
        "Clear sky": "☀️", "Mainly clear": "🌤️", "Partly cloudy": "⛅",
        "Overcast": "☁️", "Foggy": "🌫️", "Icy fog": "🌫️",
        "Light drizzle": "🌦️", "Drizzle": "🌧️", "Heavy drizzle": "🌧️",
        "Light rain": "🌦️", "Rain": "🌧️", "Heavy rain": "🌧️",
        "Light snow": "🌨️", "Snow": "❄️", "Heavy snow": "❄️",
        "Light showers": "🌦️", "Showers": "🌧️", "Heavy showers": "⛈️",
        "Thunderstorm": "⛈️", "Mixed conditions": "🌥️",
    }

    weather_icon = ""
    if weather_str:
        for desc, icon in WEATHER_ICONS.items():
            if weather_str.startswith(desc):
                weather_icon = icon
                break

    html_parts.append(
        "<table style='width:100%; border-collapse:collapse; margin-bottom:2rem;'><tr>"
        "<td style='vertical-align:bottom;'>"
        "<div style='border-left:3px solid #0F6E56; padding-left:1rem;'>"
        "<p style='font-size:12px; color:#888; margin:0 0 2px 0; letter-spacing:0.08em; text-transform:uppercase;'>Daily Update</p>"
        f"<h1 style='font-size:22px; font-weight:500; margin:0; color:#1a1a1a;'>{html_escape(subject)}</h1>"
        "</div>"
        "</td>"
        + (
            "<td style='vertical-align:bottom; text-align:right; width:140px;'>"
            f"<span style='font-size:22px; line-height:1; display:block;'>{weather_icon}</span>"
            f"<p style='font-size:12px; color:#555; margin:4px 0 0 0; white-space:nowrap;'>{html_escape(weather_str.replace(' — ', ', '))}</p>"
            "</td>"
            if weather_str else ""
        )
        + "</tr></table>"
    )

    html_parts.append(portfolio_html)

    # Workout
    html_parts.append(workout_html)

    # Quote
    html_parts.append(
        "<div style='background:#ffffff; border:0.5px solid #e0e0e0; border-radius:12px; padding:1.25rem;'>"
        "<p style='font-size:11px; color:#888; margin:0 0 10px 0; letter-spacing:0.08em; text-transform:uppercase;'>Quote</p>"
        "<blockquote style='margin:0 0 10px 0; padding-left:14px; border-left:2px solid #0F6E56;'>"
        f"<p style='font-family:Georgia,serif; font-size:15px; line-height:1.7; margin:0; font-style:italic; color:#1a1a1a;'>{html_escape(pretty)}</p>"
        "</blockquote>"
    )
    html_parts.append(
        f"<p style='font-size:13px; color:#888; margin:0;'>— {html_escape(q.author)}, <em>{html_escape(q.title)}</em></p>"
        if q.author
        else f"<p style='font-size:13px; color:#888; margin:0;'>— <em>{html_escape(q.title)}</em></p>"
    )
    html_parts.append("</div>")
    html_parts.append("</div>")

    body_html = "\n".join(html_parts)

    # ---- Send daily update ----
    send_email_smtp(
        smtp_host=smtp_host,
        smtp_port=smtp_port,
        smtp_user=smtp_user,
        smtp_pass=smtp_pass,
        mail_from=mail_from,
        mail_to=mail_to,
        subject=subject,
        body_text=body_text,
        body_html=body_html,
    )

    # ---- Send mean message ----
    ross_to = "david.mcloughlin@ucdconnect.ie"
    mean_cc = "ross.cashin@ucdconnect.ie"
    mean_text = make_insult()
    mean_subject = f"Mean Message: {today_str}"
    mean_html = (
        "<div style='font-family:Arial,sans-serif;'>"
        f"<p style='margin:0; font-size:16px;'>{html_escape(mean_text)}</p>"
        "</div>"
    )
    send_email_smtp(
        smtp_host=smtp_host,
        smtp_port=smtp_port,
        smtp_user=smtp_user,
        smtp_pass=smtp_pass,
        mail_from=mail_from,
        mail_to=ross_to,
        subject=mean_subject,
        body_text=mean_text,
        body_html=mean_html,
        mail_cc=mean_cc,
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())