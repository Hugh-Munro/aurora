import io
import os
import smtplib
import warnings
from contextlib import redirect_stdout
from datetime import date, datetime
from email.message import EmailMessage
from html import escape as html_escape
from pathlib import Path

from openpyxl import load_workbook
import portfolio_report  # <- your script: portfolio_report.py
from daily_tasks import create_daily_tasks_file
from weekly_tasks import create_weekly_tasks_file
from quote_selector import pick_random_quote

import mean_message as mean_message  # <- mean_message.py (must define make_insult())


# Suppress openpyxl warning (if any dependency triggers it)
warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
    category=UserWarning,
    module=r"openpyxl\.worksheet\._reader",
)

PASSAGES_FOLDER = Path(__file__).resolve().parent / "data" / "passages"
PDF_PATTERN = "Passages - *.pdf"
RECURSIVE = False
GAP_MULTIPLIER = 1.8

# DAILY_TASKS_TEMPLATE = Path(
#     r"C:\Users\hugom\OneDrive\Desktop\Root\Organisation\Daily Tasks\Daily_Tasks_TEMPLATE.xlsx"
# )

WEEKLY_TASKS_TEMPLATE = Path(__file__).resolve().parent / "data" / "Weekly_Tasks_TEMPLATE.xlsx"
TRAINING_PLAN_PATH = Path(__file__).resolve().parent / "data" / "Training_Plan.xlsx"

# Python weekday(): Monday=0 ... Saturday=5 ... Sunday=6
SATURDAY = 5


def load_dotenv(dotenv_path: Path) -> None:
    if not dotenv_path.exists():
        return

    for line in dotenv_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")

        if k and k not in os.environ:
            os.environ[k] = v


def file_uri(p: Path) -> str:
    """Proper file URI: file:///C:/Users/..."""
    try:
        return p.resolve().as_uri()
    except Exception:
        return "file:///" + str(p).replace("\\", "/")


def excel_uri(p: Path) -> str:
    """Excel protocol link (works in some desktop clients)."""
    return f"ms-excel:ofe|u|{file_uri(p)}"


def send_email_smtp(
    smtp_host: str,
    smtp_port: int,
    smtp_user: str,
    smtp_pass: str,
    mail_from: str,
    mail_to: str,
    subject: str,
    body_text: str,
    body_html: str | None = None,
    mail_cc: str | list[str] | None = None,  # <-- added
) -> None:
    """Plain-text + HTML alternative email sender. No attachments."""
    msg = EmailMessage()
    msg["From"] = mail_from
    msg["To"] = mail_to
    if mail_cc:
        msg["Cc"] = ", ".join(mail_cc) if isinstance(mail_cc, list) else mail_cc
    msg["Subject"] = subject

    msg.set_content(body_text)
    if body_html:
        msg.add_alternative(body_html, subtype="html")

    if smtp_port == 465:
        with smtplib.SMTP_SSL(smtp_host, smtp_port) as s:
            s.login(smtp_user, smtp_pass)
            s.send_message(msg)
    else:
        with smtplib.SMTP(smtp_host, smtp_port) as s:
            s.ehlo()
            s.starttls()
            s.login(smtp_user, smtp_pass)
            s.send_message(msg)


def prettify_quote(text: str) -> str:
    lines = [ln.lstrip().lstrip("> ").rstrip() for ln in text.splitlines()]
    return "\n".join(lines).strip()


def bullet(text: str) -> str:
    return f"• {text}"


def strip_bullet_prefix(s: str) -> str:
    return s[2:] if s.startswith("• ") else s


def _strip_done_prefix(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return s

    sl = s.lower()
    if sl.startswith("done"):
        s = s[4:].lstrip()
        if s.startswith(":") or s.startswith("-"):
            s = s[1:].lstrip()

    return s.strip()


def _norm_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "").replace("_", "")


def _cell_date_to_date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
    return None


def _find_header_row(
    ws, max_scan_rows: int = 30
) -> tuple[int, dict[str, int]] | None:
    """
    Finds a header row containing at least Week + Date + Day.
    Returns (row_index_1_based, header_map{name->col_index_1_based})
    """
    required = {"week", "date", "day"}
    known = {
        "run_km",
        "runkm",
        "run_rpe",
        "runrpe",
        "run_session",
        "runsession",
        "gym",
        "checklist",
    }

    for r in range(1, max_scan_rows + 1):
        values = [
            ws.cell(row=r, column=c).value
            for c in range(1, ws.max_column + 1)
        ]
        normed = [
            _norm_header(str(v)) if v is not None else ""
            for v in values
        ]

        if not any(normed):
            continue

        header_map: dict[str, int] = {}
        for idx, h in enumerate(normed, start=1):
            if not h:
                continue
            if h == "runkm":
                header_map["run_km"] = idx
            elif h == "runrpe":
                header_map["run_rpe"] = idx
            elif h == "runsession":
                header_map["run_session"] = idx
            elif h in {"week", "date", "day", "gym", "checklist"}:
                header_map[h] = idx
            elif h in known:
                header_map[h] = idx

        if required.issubset(header_map):
            return r, header_map

    return None


def get_today_training_plan_row(plan_path: Path, target: date) -> dict[str, str]:
    """
    Reads the Excel training plan and returns today's row as dict.
    If exact date not found, returns nearest future row and sets '_note'.
    """
    if not plan_path.exists():
        return {"_error": f"Training plan not found: {plan_path}"}

    try:
        wb = load_workbook(plan_path, data_only=True, read_only=True)
    except Exception as e:
        return {"_error": f"Could not open training plan: {e}"}

    ws = wb.active
    hdr = _find_header_row(ws)
    if hdr is None:
        return {
            "_error": (
                "Could not find header row in training plan "
                "(expected Week/Date/Day columns)."
            )
        }

    header_row, hm = hdr
    best_future = None
    found_exact = None

    def get(colname: str, row_idx: int) -> str:
        col = hm.get(colname)
        if not col:
            return ""
        v = ws.cell(row=row_idx, column=col).value
        return "" if v is None else str(v).strip()

    for r in range(header_row + 1, ws.max_row + 1):
        d = _cell_date_to_date(ws.cell(row=r, column=hm["date"]).value)
        if d is None:
            continue

        row = {
            "week": get("week", r),
            "date": d.strftime("%d/%m/%Y"),
            "day": get("day", r),
            "run_km": get("run_km", r),
            "run_rpe": get("run_rpe", r),
            "run_session": get("run_session", r),
            "gym": get("gym", r),
            "checklist": get("checklist", r),
        }

        if d == target:
            found_exact = row
            break

        if d > target:
            if best_future is None or d < best_future[0]:
                best_future = (d, row)

    if found_exact is not None:
        return found_exact

    if best_future is not None:
        d, row = best_future
        row["_note"] = (
            f"No exact match for {target.strftime('%d/%m/%Y')}; "
            f"showing next planned day {d.strftime('%d/%m/%Y')}."
        )
        return row

    return {
        "_error": (
            f"No matching dates found in plan for/after "
            f"{target.strftime('%d/%m/%Y')}."
        )
    }


def format_training_plan_row(row: dict[str, str]) -> list[str]:
    """Produces the bullet lines for the email."""
    if "_error" in row:
        return [bullet(row["_error"])]

    week = (row.get("week") or "").strip()
    dte = (row.get("date") or "").strip()
    day = (row.get("day") or "").strip()
    run_km = (row.get("run_km") or "").strip()
    run_rpe = (row.get("run_rpe") or "").strip()
    run_session = _strip_done_prefix((row.get("run_session") or "").strip())
    gym = (row.get("gym") or "").strip()
    checklist = (row.get("checklist") or "").strip()

    out: list[str] = []

    title_bits = []
    if week:
        title_bits.append(f"Week {week}")
    if day or dte:
        title_bits.append("—")
        title_bits.append(" ".join(x for x in [day, dte] if x))
    if title_bits:
        out.append(bullet(" ".join(title_bits)))

    if run_km and run_km not in {"-", ""}:
        run_line = f"Run: {run_km} km"
    else:
        run_line = "Run: (no run)"

    if run_rpe and run_rpe not in {"-", ""}:
        run_line += f" (RPE {run_rpe})"

    if run_session and run_session not in {"-", ""}:
        run_line += f" — {run_session}"

    out.append(bullet(run_line))

    if gym and gym not in {"-", ""}:
        out.append(bullet(f"Gym: {gym}"))

    if checklist and checklist not in {"-", ""}:
        out.append(bullet(f"Checklist: {checklist}"))

    note = (row.get("_note") or "").strip()
    if note:
        out.append(bullet(note))

    return out


def get_portfolio_report_pretty() -> list[str]:
    """Runs portfolio_report.main(...) and returns pretty lines."""
    import re

    def _try_float(s: str) -> float | None:
        if s is None:
            return None
        st = str(s).strip()
        if st.lower() in {"n/a", "na", ""}:
            return None
        st = st.replace(",", "")
        m = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", st)
        if not m:
            return None
        try:
            return float(m.group(0))
        except Exception:
            return None

    def _fmt_pct(x: float | None) -> str:
        return "N/A" if x is None else f"{x:.2f}%"

    def _fmt_vol_pct(x: float | None) -> str:
        return "N/A" if x is None else f"{(x * 100.0):.2f}%"

    def _fmt_2dp(x: float | None) -> str:
        return "N/A" if x is None else f"{x:.2f}"

    buf = io.StringIO()
    try:
        with redirect_stdout(buf):
            portfolio_report.main(portfolio_report.WORKBOOK_PATH)
    except Exception as e:
        return [bullet(f"(Portfolio report failed: {e})")]

    raw = buf.getvalue().splitlines()
    kv: dict[str, str] = {}

    for ln in raw:
        ln = (ln or "").strip()
        if not ln.startswith("Portfolio"):
            continue
        if ":" not in ln:
            continue
        k, v = ln.split(":", 1)
        kv[k.strip()] = v.strip()

    daily_f = _try_float(kv.get("Portfolio Daily PnL (EUR)", "N/A"))
    daily_out = "N/A" if daily_f is None else f"{daily_f:.2f}"

    return [
        bullet(f"Daily P&L (EUR): {daily_out}"),
        bullet(f"YTD Return: {_fmt_pct(_try_float(kv.get('Portfolio YTD Return (%)')))}"),
        bullet(
            f"Volatility (annualised): "
            f"{_fmt_vol_pct(_try_float(kv.get('Portfolio Vol_252d (ann)')))}"
        ),
        bullet(
            f"Sharpe: {_fmt_2dp(_try_float(kv.get('Portfolio Sharpe_252d')))}"
        ),
    ]


def main() -> int:
    here = Path(__file__).resolve().parent
    load_dotenv(here / ".env")

    smtp_host = os.environ.get("SMTP_HOST", "")
    smtp_port = int(os.environ.get("SMTP_PORT", "465"))
    smtp_user = os.environ.get("SMTP_USER", "")
    smtp_pass = os.environ.get("SMTP_PASS", "")
    mail_from = os.environ.get("MAIL_FROM", smtp_user)
    mail_to = os.environ.get("MAIL_TO", "")

    missing = [
        k
        for k in [
            "SMTP_HOST",
            "SMTP_PORT",
            "SMTP_USER",
            "SMTP_PASS",
            "MAIL_FROM",
            "MAIL_TO",
        ]
        if not os.environ.get(k)
    ]
    if missing:
        raise SystemExit(f"Missing settings in .env: {', '.join(missing)}")

    # ---- Portfolio Report ----
    portfolio_lines = get_portfolio_report_pretty()

    # ---- Weekly Tasks: ONLY run/include on Saturdays ----
    is_saturday = date.today().weekday() == SATURDAY
    weekly_path = None
    include_weekly = False
    weekly_file_uri = None
    weekly_excel_uri = None

    if is_saturday:
        weekly_result = create_weekly_tasks_file(WEEKLY_TASKS_TEMPLATE)
        weekly_path = weekly_result.path

        if (
            weekly_path.exists()
            and weekly_path.is_file()
            and weekly_path.suffix.lower() == ".xlsx"
            and weekly_path.name.startswith("Weekly_Tasks_")
            and weekly_path.name != WEEKLY_TASKS_TEMPLATE.name
        ):
            include_weekly = True
            weekly_file_uri = file_uri(weekly_path)
            weekly_excel_uri = excel_uri(weekly_path)

    # ---- Workout of the Day ----
    plan_row = get_today_training_plan_row(TRAINING_PLAN_PATH, date.today())
    workout_lines = format_training_plan_row(plan_row)

    # ---- Quote selection ----
    q = pick_random_quote(
        PASSAGES_FOLDER,
        pattern=PDF_PATTERN,
        recursive=RECURSIVE,
        gap_multiplier=GAP_MULTIPLIER,
    )
    pretty = prettify_quote(q.text)

    # ---- Build daily update email ----
    today_str = date.today().strftime("%a %d %b %Y")
    subject = f"Daily Update: {today_str}"

    lines: list[str] = []
    lines.append(subject)
    lines.append("")

    lines.append("Portfolio Report")
    lines.extend(portfolio_lines)
    lines.append("")

    if include_weekly and weekly_path and weekly_file_uri:
        lines.append("Weekly Tasks")
        lines.append(bullet(str(weekly_path)))
        lines.append(bullet(weekly_file_uri))
        lines.append("")

    lines.append("Workout")
    lines.extend(workout_lines)
    lines.append("")

    lines.append("Quote")
    lines.append(f"“{pretty}”")
    lines.append(f"— {q.author}, {q.title}" if q.author else f"— {q.title}")

    body_text = "\n".join(lines)

    html_parts: list[str] = []
    html_parts.append("<div style='font-family:Segoe UI, Arial, sans-serif;'>")
    html_parts.append(
        f"<h1 style='margin:0 0 16px 0; font-size:25px; line-height:1.2;'>"
        f"{html_escape(subject)}</h1>"
    )

    html_parts.append(
        "<h2 style='margin:18px 0 8px 0; font-size:18px; font-weight:700;'>"
        "Portfolio Report</h2>"
    )
    html_parts.append("<ul>")
    for pl in portfolio_lines:
        html_parts.append(f"<li>{html_escape(strip_bullet_prefix(pl))}</li>")
    html_parts.append("</ul>")

    if include_weekly and weekly_excel_uri and weekly_path:
        html_parts.append(
            "<h2 style='margin:18px 0 8px 0; font-size:18px; font-weight:700;'>"
            "Weekly Tasks</h2>"
        )
        html_parts.append("<ul>")
        html_parts.append(
            f"<li><a href='{html_escape(weekly_excel_uri)}'>"
            f"{html_escape(str(weekly_path))}</a></li>"
        )
        html_parts.append("</ul>")

    html_parts.append(
        "<h2 style='margin:18px 0 8px 0; font-size:18px; font-weight:700;'>Workout</h2>"
    )
    html_parts.append("<ul>")
    for wl in workout_lines:
        html_parts.append(f"<li>{html_escape(strip_bullet_prefix(wl))}</li>")
    html_parts.append("</ul>")

    html_parts.append(
        "<h2 style='margin:18px 0 8px 0; font-size:18px; font-weight:700;'>Quote</h2>"
    )
    html_parts.append(
        "<blockquote style='margin:0 0 8px 0; padding-left:12px; "
        "border-left:3px solid #ddd;'>"
        f"{html_escape(pretty)}</blockquote>"
    )
    html_parts.append(
        f"<div>— {html_escape(q.author)}, {html_escape(q.title)}</div>"
        if q.author
        else f"<div>— {html_escape(q.title)}</div>"
    )
    html_parts.append("</div>")

    body_html = "\n".join(html_parts)

    # --- Send YOUR daily update to MAIL_TO (NO CC) ---
    send_email_smtp(
        smtp_host=smtp_host,
        smtp_port=smtp_port,
        smtp_user=smtp_user,
        smtp_pass=smtp_pass,
        mail_from=mail_from,
        mail_to=mail_to,
        subject=subject,
        body_text=body_text,
        body_html=body_html,
    )

    # --- Send ONLY the mean message to Ross (WITH CC) ---
    ross_to = "david.mcloughlin@ucdconnect.ie"
    mean_cc = "ross.cashin@ucdconnect.ie"  # <-- set your CC email here (or make it a list)
    mean_text = mean_message.make_insult()  # mean_message.py must define this
    mean_subject = f"Mean Message: {today_str}"
    mean_html = (
        "<div style='font-family:Segoe UI, Arial, sans-serif;'>"
        f"<p style='margin:0; font-size:16px;'>{html_escape(mean_text)}</p>"
        "</div>"
    )

    send_email_smtp(
        smtp_host=smtp_host,
        smtp_port=smtp_port,
        smtp_user=smtp_user,
        smtp_pass=smtp_pass,
        mail_from=mail_from,
        mail_to=ross_to,
        subject=mean_subject,
        body_text=mean_text,   # ONLY the mean message
        body_html=mean_html,   # ONLY the mean message
        mail_cc=mean_cc,       # <-- CC only here
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())