import os
import re
import statistics
from pathlib import Path
import pdfplumber
import openpyxl
from openpyxl import load_workbook


# ------------------ PATHS ------------------

EXCEL_PATH = r"C:\Users\hugom\OneDrive\Desktop\Root\Personal\Arts\Books.xlsx"
PDF_FOLDER = r"C:\Users\hugom\OneDrive\Desktop\Root\Personal\Arts\Passages"


# ------------------ HELPERS ------------------

def extract_book_title(filename: str) -> str:
    """
    Extract book title from filename like:
    'Passages - White Nights.pdf'
    """
    m = re.match(r"Passages\s*-\s*(.+)\.pdf", filename, re.I)
    return m.group(1).strip() if m else filename.replace(".pdf", "").strip()


def _join_lines(lines):
    """
    Join wrapped PDF lines into one clean paragraph.
    """
    text = " ".join(l.strip() for l in lines if l.strip())
    text = re.sub(r"\s+", " ", text)
    return text.strip()


# ------------------ CORE EXTRACTION ------------------

def extract_quotes_from_pdf(pdf_path: str) -> list[str]:
    """
    Extract quotes by detecting vertical spacing between text blocks.
    This is the ONLY reliable method for your PDFs.
    """
    quotes: list[str] = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            lines = page.extract_text_lines(strip=True) or []
            if not lines:
                continue

            # Sort top → bottom
            lines.sort(key=lambda l: (l["top"], l["x0"]))

            # Measure vertical gaps
            gaps = []
            for i in range(1, len(lines)):
                gaps.append(lines[i]["top"] - lines[i - 1]["bottom"])

            median_gap = statistics.median(gaps) if gaps else 0
            threshold = max(median_gap * 1.8, 3)

            buffer = [lines[0]["text"]]

            for i in range(1, len(lines)):
                gap = lines[i]["top"] - lines[i - 1]["bottom"]

                if gap > threshold:
                    paragraph = _join_lines(buffer)
                    if len(paragraph) > 30:
                        quotes.append(paragraph)
                    buffer = [lines[i]["text"]]
                else:
                    buffer.append(lines[i]["text"])

            # Flush final buffer
            paragraph = _join_lines(buffer)
            if len(paragraph) > 30:
                quotes.append(paragraph)

    # Remove headers if they appear
    cleaned = []
    for q in quotes:
        q = re.sub(r"^Passages\s*-\s*[^\n]+", "", q, flags=re.I).strip()
        if q:
            cleaned.append(q)

    return cleaned


# ------------------ EXCEL ------------------

def add_quotes_to_excel(excel_path: str, all_quotes: list[dict]) -> None:
    """
    Append extracted quotes to the Quotes sheet.
    """
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    if "Quotes" in wb.sheetnames:
        ws = wb["Quotes"]
    else:
        ws = wb.create_sheet("Quotes")
        ws.append(["Book Title", "Quote", "Topic/Tags", "Notes"])

    for q in all_quotes:
        ws.append([
            q["book_title"],
            q["quote"],
        ])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 30

    wb.save(excel_path)
    print(f"\n✓ Saved {len(all_quotes)} quotes to {excel_path}")


# ------------------ MAIN ------------------

def main():
    print("Starting PDF quote extraction")
    print(f"PDF folder : {PDF_FOLDER}")
    print(f"Excel file: {EXCEL_PATH}")
    print("-" * 60)

    pdfs = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith(".pdf")]

    if not pdfs:
        print("No PDFs found.")
        return

    all_quotes = []

    for pdf in pdfs:
        pdf_path = os.path.join(PDF_FOLDER, pdf)
        title = extract_book_title(pdf)

        print(f"Processing: {title}")
        quotes = extract_quotes_from_pdf(pdf_path)

        for q in quotes:
            all_quotes.append({
                "book_title": title,
                "quote": q
            })

        print(f"  → {len(quotes)} quotes")

    if all_quotes:
        print("-" * 60)
        print(f"Total quotes extracted: {len(all_quotes)}")
        add_quotes_to_excel(EXCEL_PATH, all_quotes)
        print("\n✓ Done.")
    else:
        print("No quotes extracted.")


if __name__ == "__main__":
    main()
