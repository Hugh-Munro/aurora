from __future__ import annotations

import warnings
import math
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook


# ============================================================
# SUPPRESS WARNINGS
# ============================================================

warnings.filterwarnings("ignore", category=UserWarning, module=r"openpyxl")


# ============================================================
# CONFIG
# ============================================================

DAILY_TASKS_FOLDER = Path(
    r"C:\Users\hugom\OneDrive\Desktop\Root\Organisation\Daily Tasks"
)

WEEKLY_TASKS_FOLDER = Path(
    r"C:\Users\hugom\OneDrive\Desktop\Root\Organisation\Weekly Tasks"
)

WEIGHT_FILE = Path(
    r"C:\Users\hugom\OneDrive\Desktop\Root\Personal\Fitness\Diet\Weight.xlsx"
)

PORTFOLIO_FILE = Path(
    r"C:\Users\hugom\OneDrive\Desktop\Root\Personal\Projects\Portfolio\Portfolio.xlsx"
)

MOVIES_FILE = Path(
    r"C:\Users\hugom\OneDrive\Desktop\Root\Personal\Arts\movie_dataset.xlsx"
)

LIKED_SONGS_FILE = Path(
    r"C:\Users\hugom\OneDrive\Desktop\Root\Personal\Arts\Music\Liked_Songs.csv"
)

BOOKS_FILE = Path(
    r"C:\Users\hugom\OneDrive\Desktop\Root\Personal\Arts\book_dataset.xlsx"
)

BOOKS_SHEET = "Read Books"

STUDY_FILE = Path(
    r"C:\Users\hugom\OneDrive\Desktop\Root\Career\Undergrad\Study_Log.xlsx"
)

WORKOUTS_FILE = Path(
    r"C:\Users\hugom\OneDrive\Desktop\Root\Personal\Fitness\Sports\Gym\workouts.csv"
)

RUNNING_FILE = Path(
    r"C:\Users\hugom\OneDrive\Desktop\Root\Personal\Fitness\Sports\Running\Activities.csv"
)

YES_VALUE = "yes"
NO_VALUE = "no"


# ============================================================
# SHARED HELPERS
# ============================================================

def _is_same_month(d: date, year: int, month: int) -> bool:
    return d.year == year and d.month == month


def _extract_date_from_filename(name: str) -> date | None:
    try:
        ymd = name.split("_")[-1].split(".")[0]
        return date(int(ymd[0:4]), int(ymd[4:6]), int(ymd[6:8]))
    except Exception:
        return None


def _parse_excel_date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except Exception:
                pass
    return None


def _parse_iso_datetime(v) -> date | None:
    if v is None:
        return None
    try:
        return datetime.fromisoformat(str(v).replace("Z", "+00:00")).date()
    except Exception:
        return None


def _parse_workout_datetime(v) -> datetime | None:
    try:
        return datetime.strptime(str(v).strip(), "%d %b %Y, %H:%M")
    except Exception:
        return None


def _parse_running_datetime(v) -> datetime | None:
    """
    Parses Garmin format:
      05/01/2026 15:07
    """
    try:
        return datetime.strptime(str(v).strip(), "%d/%m/%Y %H:%M")
    except Exception:
        return None


# ============================================================
# TASKS
# ============================================================

def _iter_task_files(folder: Path, year: int, month: int) -> Iterable[Path]:
    for p in folder.glob("*.xlsx"):
        d = _extract_date_from_filename(p.name)
        if d and _is_same_month(d, year, month):
            yield p


def _count_tasks_in_file(path: Path) -> tuple[int, int]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    completed = 0
    relevant = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        for cell in row:
            if cell is None:
                continue
            v = str(cell).strip().lower()
            if v == YES_VALUE:
                completed += 1
                relevant += 1
            elif v == NO_VALUE:
                relevant += 1

    return completed, relevant


def calculate_monthly_completion(year: int, month: int) -> float:
    done = 0
    total = 0

    for folder in (DAILY_TASKS_FOLDER, WEEKLY_TASKS_FOLDER):
        for f in _iter_task_files(folder, year, month):
            d, t = _count_tasks_in_file(f)
            done += d
            total += t

    return 0.0 if total == 0 else (done / total) * 100.0


# ============================================================
# WEIGHT
# ============================================================

def get_monthly_weight_change(year: int, month: int):
    if not WEIGHT_FILE.exists():
        return None, None, None

    wb = load_workbook(WEIGHT_FILE, data_only=True, read_only=True)
    ws = wb.active

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None or r[1] is None:
            continue
        d = _parse_excel_date(r[0])
        if d and _is_same_month(d, year, month):
            rows.append((d, float(r[1])))

    if not rows:
        return None, None, None

    rows.sort(key=lambda x: x[0])
    return rows[0][1], rows[-1][1], rows[-1][1] - rows[0][1]


# ============================================================
# PORTFOLIO — MONTHLY PNL
# ============================================================

def get_monthly_portfolio_pnl(year: int, month: int) -> float | None:
    if not PORTFOLIO_FILE.exists():
        return None

    df = pd.read_excel(PORTFOLIO_FILE)
    df = df[["YF_Ticker", "Shares"]].dropna()
    df = df[df["Shares"] != 0]

    if df.empty:
        return None

    tickers = df["YF_Ticker"].tolist()
    shares = df.set_index("YF_Ticker")["Shares"].astype(float)

    start = pd.Timestamp(year=year, month=month, day=1)
    end = start + pd.offsets.MonthEnd(1)

    prices = yf.download(
        tickers,
        start=start - pd.Timedelta(days=5),
        end=end + pd.Timedelta(days=5),
        progress=False,
        group_by="ticker",
        auto_adjust=False,
    )

    total_start = 0.0
    total_end = 0.0

    for t in tickers:
        try:
            s = prices[t]["Close"].dropna()
            s_month = s[(s.index >= start) & (s.index <= end)]
            if s_month.empty:
                continue
            total_start += s_month.iloc[0] * shares[t]
            total_end += s_month.iloc[-1] * shares[t]
        except Exception:
            continue

    return None if total_start == 0 else (total_end - total_start)


# ============================================================
# MOVIES
# ============================================================

def get_movies_watched_in_month(year: int, month: int) -> int:
    if not MOVIES_FILE.exists():
        return 0

    wb = load_workbook(MOVIES_FILE, data_only=True, read_only=True)
    ws = wb.active

    header = [str(c).strip() if c else "" for c in next(ws.iter_rows(values_only=True))]
    try:
        date_col = header.index("Date Watched")
    except ValueError:
        return 0

    return sum(
        1
        for row in ws.iter_rows(min_row=2, values_only=True)
        if (
            (d := _parse_excel_date(row[date_col] if date_col < len(row) else None))
            and _is_same_month(d, year, month)
        )
    )


# ============================================================
# MUSIC — LIKED SONGS
# ============================================================

def get_liked_songs_added_in_month(year: int, month: int) -> int:
    if not LIKED_SONGS_FILE.exists():
        return 0

    df = pd.read_csv(LIKED_SONGS_FILE)
    if "Added At" not in df.columns:
        return 0

    return sum(
        1
        for v in df["Added At"]
        if (d := _parse_iso_datetime(v)) and _is_same_month(d, year, month)
    )


# ============================================================
# BOOKS
# ============================================================

def get_books_read_in_month(year: int, month: int) -> int:
    if not BOOKS_FILE.exists():
        return 0

    wb = load_workbook(BOOKS_FILE, data_only=True, read_only=True)
    try:
        ws = wb[BOOKS_SHEET]
    except KeyError:
        return 0

    header = [str(c).strip() if c else "" for c in next(ws.iter_rows(values_only=True))]
    try:
        date_col = header.index("Read Date")
    except ValueError:
        return 0

    return sum(
        1
        for row in ws.iter_rows(min_row=2, values_only=True)
        if (
            (d := _parse_excel_date(row[date_col] if date_col < len(row) else None))
            and _is_same_month(d, year, month)
        )
    )


# ============================================================
# STUDY
# ============================================================

def get_monthly_study_hours(year: int, month: int) -> float:
    if not STUDY_FILE.exists():
        return 0.0

    wb = load_workbook(STUDY_FILE, data_only=True, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows or rows[0][0] != "Date":
        return 0.0

    total = 0.0

    for row in rows[1:]:
        d = _parse_excel_date(row[0])
        if d and _is_same_month(d, year, month):
            for v in row[1:]:
                try:
                    if v is not None:
                        total += float(v)
                except Exception:
                    continue

    return total


# ============================================================
# WORKOUTS
# ============================================================

def get_monthly_workouts(year: int, month: int) -> tuple[int, float]:
    if not WORKOUTS_FILE.exists():
        return 0, 0.0

    df = pd.read_csv(WORKOUTS_FILE)
    df["parsed_start"] = df["start_time"].apply(_parse_workout_datetime)
    df = df[df["parsed_start"].notna()]
    df = df[df["parsed_start"].apply(lambda d: _is_same_month(d.date(), year, month))]

    sessions = df["start_time"].nunique()

    total_kg = 0.0
    for _, r in df.iterrows():
        w, reps = r.get("weight_kg"), r.get("reps")
        if w is None or reps is None:
            continue
        try:
            w, reps = float(w), float(reps)
        except Exception:
            continue
        if math.isnan(w) or math.isnan(reps):
            continue
        total_kg += w * reps

    return sessions, total_kg


# ============================================================
# RUNNING
# ============================================================

def get_monthly_running(year: int, month: int) -> tuple[int, float]:
    """
    Returns:
      (number_of_runs, total_km)
    """
    if not RUNNING_FILE.exists():
        return 0, 0.0

    df = pd.read_csv(RUNNING_FILE)

    if "Date" not in df.columns or "Distance" not in df.columns:
        return 0, 0.0

    df["parsed_date"] = df["Date"].apply(_parse_running_datetime)
    df = df[df["parsed_date"].notna()]
    df = df[df["parsed_date"].apply(lambda d: _is_same_month(d.date(), year, month))]

    if df.empty:
        return 0, 0.0

    runs = len(df)

    total_km = 0.0
    for v in df["Distance"]:
        try:
            total_km += float(str(v).replace(",", ""))
        except Exception:
            continue

    return runs, total_km


# ============================================================
# CLI
# ============================================================

def main() -> None:
    today = date.today()

    task_pct = calculate_monthly_completion(today.year, today.month)
    w0, w1, wd = get_monthly_weight_change(today.year, today.month)
    pnl = get_monthly_portfolio_pnl(today.year, today.month)
    movies = get_movies_watched_in_month(today.year, today.month)
    songs = get_liked_songs_added_in_month(today.year, today.month)
    books = get_books_read_in_month(today.year, today.month)
    study = get_monthly_study_hours(today.year, today.month)
    sessions, kg = get_monthly_workouts(today.year, today.month)
    runs, km = get_monthly_running(today.year, today.month)

    print(f"Monthly task completion: {task_pct:.2f}%")

    if wd is None:
        print("Monthly weight change: insufficient data")
    else:
        print(f"Weight change: {w0:.1f} kg → {w1:.1f} kg ({wd:+.1f} kg)")

    if pnl is None:
        print("Monthly portfolio PnL: insufficient data")
    else:
        print(f"Monthly portfolio PnL: €{pnl:,.2f}")

    print(f"Movies watched this month: {movies}")
    print(f"New liked songs this month: {songs}")
    print(f"Books read this month: {books}")
    print(f"Study hours this month: {study:.1f} h")
    print(f"Gym sessions this month: {sessions}")
    print(f"Total weight lifted this month: {kg:,.0f} kg")
    print(f"Runs this month: {runs}")
    print(f"Distance run this month: {km:.2f} km")


if __name__ == "__main__":
    main()
